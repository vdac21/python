#encoding:UTF_8
import sys
from PyQt5.QtWidgets import *
#from PyQt5.QtGui import QDateEdit
from openpyxl import load_workbook

class Window(QWidget):
    
    def _init_(self):
        super(Main, self)._init_()
        self.setWindowTitle("Load Excel data to QTable")
        
        

    def __init__(self):
        super().__init__()
        self.setWindowTitle("PYTHON")
        self.resize(400, 300)
        # Create a top-level layout
        layout = QVBoxLayout()
        self.setLayout(layout)
        # Create the tab widget with two tabs
        tabs = QTabWidget()
        tabs.addTab(self.ADD_USER(), "ADD_USER")
        tabs.addTab(self.LIST_USER(), "LIST_USER")
        layout.addWidget(tabs)

    def ADD_USER(self):
        """Create the General page UI."""
        tab1_user_details = QWidget()
        layout_gender = QHBoxLayout()
        layout1=QFormLayout()
        #user components for tab1
        self.obj_le_fname = QLineEdit()
        self.obj_le_lname = QLineEdit()
        self.obj_le_email = QLineEdit()
        self.obj_le_dob = QLineEdit()
        self.obj_bt_submit = QPushButton("SUBMIT")
        self.obj_rbt_male = QRadioButton("Male")
        self.obj_rbt_female = QRadioButton("Female")

        #adding radio buttons to Hbox layout
        layout_gender.addWidget(self.obj_rbt_male)
        layout_gender.addWidget(self.obj_rbt_female)

        #adding all components to the form layout
        layout1.addRow("FIRST NAME",self.obj_le_fname)
        layout1.addRow("LAST NAME",self.obj_le_lname)
        layout1.addRow("EMAIL",self.obj_le_email)
        layout1.addRow("DOB",self.obj_le_dob)       
        layout1.addRow(QLabel("Gender"),layout_gender)        
        layout1.addRow(self.obj_bt_submit)

        #adding form layout to tab1_widget
        tab1_user_details.setLayout(layout1)
        self.obj_bt_submit.clicked.connect(self.submit_action)
        return tab1_user_details
    def load_data(self):
         path = "https://d.docs.live.net/cd64eef9b8ea2994/pyqt5%20data.xlsx"
         workbook = openpyxl.load_workbook(path)
         sheet = workbook.active

    def add_to_data(self):
        wb = load_workbook(filename =("C:\\Users\\BHGYALAKSHMI\\OneDrive\\Desktop\\python practing\\mu ui xlsx"))
        ws = wb['sheet1']

    def submit_action(self):
        fname = self.obj_le_fname.text()
        print(fname)
        lname = self.obj_le_lname.text()
        print(lname)
        email = self.obj_le_email.text()
        print(email)
        dob = self.obj_le_dob.text()
        print(dob)
      
	
    def maleselected(self, selected):
       if selected:
           self.label.setText("MALE")
			
    def femaleselected(self, selected):
       if selected:
           self.label.setText("FEMALE")	
			
    def retranslateUi(self, MainWindow):
	_translate = QtCore.QCoreApplication.translate

	MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
	self.radioButton_male.setText(_translate("MainWindow", "Male"))
	self.label.setText(_translate("MainWindow", "Select your gender:"))
	self.radioButton_female.setText(_translate("MainWindow", "Female"))

# Driver Code
if __name__ == "__main__":
	app = QtWidgets.QApplication(sys.argv)

	MainWindow = QtWidgets.QMainWindow()
	ui = Ui_MainWindow()
	ui.setupUi(MainWindow)
	MainWindow.show()
	sys.exit(app.exec_())
  
        #male = self.obj_rbt_male.text()
        #print(male)
        female = self.obj_rbt_female.text()
        print(female)

    def LIST_USER(self):
        """Create the Network page UI."""
        listuserTab = QWidget()
        
        
        return listuserTab

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Window()
    window.show()
    sys.exit(app.exec_())
